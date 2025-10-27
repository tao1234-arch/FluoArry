import cv2
import numpy as np
import openpyxl
from skimage.restoration import denoise_wavelet
from deap import algorithms, base, creator, tools
import random
from minisom import MiniSom
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import matplotlib
import tkinter as tk
from tkinter import filedialog
from openpyxl.utils import get_column_letter

matplotlib.use('TkAgg')  # 使用 TkAgg 后端
import matplotlib.pyplot as plt


class MicroarrayGridding:
    def __init__(self, image_path):
        self.image = cv2.imread(image_path)
        self.r_channel = self.image[:, :, 2]  # 红色通道
        self.g_channel = self.image[:, :, 1]  # 绿色通道
        self.grid_size = self.detect_grid_size()  # 动态检测网格大小
        self.image_path = image_path  # 保存图像路径

    # 通过行列投影和峰值检测去确定网格线的数量
    def detect_grid_size(self):
        """自动检测微阵列的行列数"""
        # 确保使用灰度图进行投影计算
        gray = cv2.cvtColor(self.image, cv2.COLOR_BGR2GRAY) if len(self.image.shape) == 3 else self.image

        # 计算行列投影
        row_sum = np.sum(gray, axis=1)  # 垂直投影（每行的总和）
        col_sum = np.sum(gray, axis=0)  # 水平投影（每列的总和）

        # 平滑投影曲线以减少噪声影响
        def smooth(x, window_len=11, window='hanning'):
            if x.ndim != 1:
                raise ValueError("smooth only accepts 1 dimension arrays.")
            if x.size < window_len:
                raise ValueError("Input vector needs to be bigger than window size.")
            if window_len < 3:
                return x
            if not window in ['flat', 'hanning', 'hamming', 'bartlett', 'blackman']:
                raise ValueError("Window is on of 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'")

            s = np.r_[x[window_len - 1:0:-1], x, x[-2:-window_len - 1:-1]]
            if window == 'flat':  # moving average
                w = np.ones(window_len, 'd')
            else:
                w = eval('np.' + window + '(window_len)')

            y = np.convolve(w / w.sum(), s, mode='valid')
            return y[int(window_len / 2 - 1):int(-window_len / 2)]

        # 平滑投影曲线
        row_sum_smooth = smooth(row_sum, window_len=25)
        col_sum_smooth = smooth(col_sum, window_len=25)

        # 寻找行和列的峰值点（信号强的区域，对应微阵列点）
        row_peaks, _ = find_peaks(row_sum_smooth, distance=10, prominence=1000)
        col_peaks, _ = find_peaks(col_sum_smooth, distance=10, prominence=1000)

        # 计算行列数（峰值数量即为点的数量）
        rows_count = len(row_peaks)
        cols_count = len(col_peaks)

        print(f"自动检测到的网格大小: {rows_count}行 x {cols_count}列")
        return (rows_count, cols_count)

    def calculate_grid_lines(self):
        """通过行列投影的谷值点确定网格线"""
        # 确保使用灰度图进行投影计算
        gray = cv2.cvtColor(self.image, cv2.COLOR_BGR2GRAY) if len(self.image.shape) == 3 else self.image

        # 计算行列投影
        row_sum = np.sum(gray, axis=1)  # 垂直投影（每行的总和）
        col_sum = np.sum(gray, axis=0)  # 水平投影（每列的总和）

        # 平滑投影曲线以减少噪声影响
        def smooth(x, window_len=11, window='hanning'):
            if x.ndim != 1:
                raise ValueError("smooth only accepts 1 dimension arrays.")
            if x.size < window_len:
                raise ValueError("Input vector needs to be bigger than window size.")
            if window_len < 3:
                return x
            if not window in ['flat', 'hanning', 'hamming', 'bartlett', 'blackman']:
                raise ValueError("Window is on of 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'")

            s = np.r_[x[window_len - 1:0:-1], x, x[-2:-window_len - 1:-1]]
            if window == 'flat':  # moving average
                w = np.ones(window_len, 'd')
            else:
                w = eval('np.' + window + '(window_len)')

            y = np.convolve(w / w.sum(), s, mode='valid')
            return y[int(window_len / 2 - 1):int(-window_len / 2)]

        # 平滑投影曲线
        row_sum_smooth = smooth(row_sum, window_len=25)
        col_sum_smooth = smooth(col_sum, window_len=25)

        # 寻找行和列的谷值点（局部最小值）
        # 注意：我们寻找的是投影曲线的谷值点（信号弱的区域），对应网格线位置
        row_valleys, _ = find_peaks(-row_sum_smooth, distance=10, prominence=1000)
        col_valleys, _ = find_peaks(-col_sum_smooth, distance=10, prominence=1000)

        # 如果找到的谷值点数量不足，则使用等距网格作为后备
        if len(row_valleys) < self.grid_size[0] - 1:
            row_valleys = np.linspace(10, self.image.shape[0] - 10, self.grid_size[0] - 1, dtype=int)

        if len(col_valleys) < self.grid_size[1] - 1:
            col_valleys = np.linspace(10, self.image.shape[1] - 10, self.grid_size[1] - 1, dtype=int)

        # 添加图像边界作为网格线的起点和终点
        rows = np.concatenate([[0], row_valleys, [self.image.shape[0] - 1]])
        cols = np.concatenate([[0], col_valleys, [self.image.shape[1] - 1]])

        return rows, cols

    # 通过高亮区域去调整网格线
    def local_som_adjustment(self, grid_cells, rows, cols):
        """
        使用基于密度分析的方法调整网格线，确保每个荧光点位于网格单元内
        """
        # 第一步：提取每个网格的质心和密度特征
        cell_features = []
        for i in range(len(rows) - 1):
            for j in range(len(cols) - 1):
                idx = i * (len(cols) - 1) + j
                cell = grid_cells[idx]

                # 计算荧光点分布特征
                if cell.size == 0:
                    # 空单元格处理
                    features = {
                        'has_points': False,
                        'density': 0,
                        'centroid': (0.5, 0.5),  # 中心位置
                        'coverage': 0
                    }
                else:
                    # 计算自适应阈值，通过阈值分割，快速识别高亮区域，假设这些区域是荧光点
                    threshold = max(50, np.percentile(cell, 80))
                    binary = cell > threshold

                    # 计算荧光点坐标
                    y_coords, x_coords = np.nonzero(binary)

                    if len(y_coords) > 0:
                        # 计算质心（归一化到单元格内），计算所有高亮像素的平均值
                        cy = np.mean(y_coords) / cell.shape[0]
                        cx = np.mean(x_coords) / cell.shape[1]

                        # 计算密度和覆盖率
                        density = np.sum(binary) / binary.size
                        coverage = len(y_coords) / (cell.shape[0] * cell.shape[1])

                        features = {
                            'has_points': True,
                            'density': density,
                            'centroid': (cy, cx),
                            'coverage': coverage,
                            'intensity': np.mean(cell[binary])
                        }
                    else:
                        # 无显著荧光点
                        features = {
                            'has_points': False,
                            'density': 0,
                            'centroid': (0.5, 0.5),
                            'coverage': 0
                        }

                    cell_features.append((i, j, features))

        # 第二步：基于荧光点分布调整网格线
        row_adjustments = np.zeros(len(rows))
        col_adjustments = np.zeros(len(cols))

        # 计算调整量
        for i, j, features in cell_features:
            if not features['has_points']:
                continue

            cy, cx = features['centroid']
            density = features['density']
            intensity = features.get('intensity', 100) / 255.0

            # 计算相对偏移 (0.5为中心基准)
            dy_rel = cy - 0.5
            dx_rel = cx - 0.5

            # 获取当前网格大小
            row_interval = rows[i + 1] - rows[i]
            col_interval = cols[j + 1] - cols[j]

            # 计算调整强度（基于密度和亮度）
            adjustment_strength = min(1.0, max(0.3, density * intensity * 2))

            # 计算实际调整量（限制在网格大小的20%以内）
            max_adjust = 0.4
            dy = np.clip(dy_rel * row_interval * adjustment_strength,
                         -max_adjust * row_interval, max_adjust * row_interval)
            dx = np.clip(dx_rel * col_interval * adjustment_strength,
                         -max_adjust * col_interval, max_adjust * col_interval)

            # 累积调整量（使用加权方式，中心点权重更高）
            # 调整网格的平滑度和局部性
            # weight = 0.7  # 中心点权重
            # 高密度区域使用较低权重（更平滑），低密度区域使用较高权重（更精确）
            weight = 0.4 + min(0.3, density * 2)
            # 相邻点权重
            row_adjustments[i] -= dy * (1 - weight)
            row_adjustments[i + 1] += dy
            # 修复行调整的边界条件：确保i+2 < 行调整数组长度
            if (i + 2) < len(row_adjustments):
                row_adjustments[i + 2] -= dy * (1 - weight)

            col_adjustments[j] -= dx * (1 - weight)
            col_adjustments[j + 1] += dx
            # 修复列调整的边界条件：确保j+2 < 列调整数组长度
            if (j + 2) < len(col_adjustments):
                col_adjustments[j + 2] -= dx * (1 - weight)

        # 第三步：平滑调整量并应用约束，消除局部调整引起的网格线抖动，保持整体规律性
        def smooth_adjustments(adjustments, window_size=3):
            smoothed = adjustments.copy()
            half_window = window_size // 2

            for i in range(half_window, len(adjustments) - half_window):
                # 加权平均，中心权重更高
                weights = np.ones(window_size)
                weights[half_window] = 2.0  # 中心权重加倍
                weights /= np.sum(weights)

                # 应用加权平均
                smoothed[i] = np.sum(adjustments[i - half_window:i + half_window + 1] * weights)

            return smoothed

        # 平滑调整量
        row_adjustments = smooth_adjustments(row_adjustments)
        col_adjustments = smooth_adjustments(col_adjustments)

        # 应用调整量，生成最终网格线
        adjusted_rows = rows.copy().astype(float)
        adjusted_cols = cols.copy().astype(float)

        # 应用调整量，确保网格线顺序不变且间距合理
        min_spacing = 10  # 最小网格间距
        for i in range(1, len(adjusted_rows) - 1):
            adjusted_rows[i] += row_adjustments[i]
            # 确保网格线不交叉且保持最小间距
            adjusted_rows[i] = max(adjusted_rows[i - 1] + min_spacing,
                                   min(adjusted_rows[i], adjusted_rows[i + 1] - min_spacing))

        for j in range(1, len(adjusted_cols) - 1):
            adjusted_cols[j] += col_adjustments[j]
            # 确保网格线不交叉且保持最小间距
            adjusted_cols[j] = max(adjusted_cols[j - 1] + min_spacing,
                                   min(adjusted_cols[j], adjusted_cols[j + 1] - min_spacing))

        return adjusted_rows.astype(int), adjusted_cols.astype(int)

    # 网格划分的主流程
    def process(self):
        """执行网格划分和调整的主流程"""
        rows, cols = self.calculate_grid_lines()

        # 使用红色通道创建网格单元
        grid_cells = [self.r_channel[rows[i]:rows[i + 1], cols[j]:cols[j + 1]]
                      for i in range(len(rows) - 1) for j in range(len(cols) - 1)]

        # 对网格进行SOM局部动态调整
        adjusted_rows, adjusted_cols = self.local_som_adjustment(grid_cells, rows, cols)

        return adjusted_rows, adjusted_cols

    def save_grid_to_excel(self, adjusted_rows, adjusted_cols, output_path=None):
        """
        将网格位置信息保存到Excel表格
        :param adjusted_rows: 调整后的行网格线
        :param adjusted_cols: 调整后的列网格线
        :param output_path: Excel文件保存路径，默认为原图路径加_grid_info.xlsx
        """
        if output_path is None:
            # 生成默认输出路径，在原图路径基础上修改
            base_name = self.image_path.split('.')[0]
            output_path = f"{base_name}_grid_info.xlsx"

        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active

        # 设置表头
        ws['A1'] = "网格编号"
        ws['B1'] = "行索引"
        ws['C1'] = "列索引"
        ws['D1'] = "左上角X坐标"
        ws['E1'] = "左上角Y坐标"
        ws['F1'] = "右下角X坐标"
        ws['G1'] = "右下角Y坐标"
        ws['H1'] = "网格宽度"
        ws['I1'] = "网格高度"

        # 填充网格信息
        grid_count = 1
        for i in range(len(adjusted_rows) - 1):
            for j in range(len(adjusted_cols) - 1):
                # 计算网格坐标
                x1 = adjusted_cols[j]
                y1 = adjusted_rows[i]
                x2 = adjusted_cols[j + 1]
                y2 = adjusted_rows[i + 1]

                # 计算网格尺寸
                width = x2 - x1
                height = y2 - y1

                # 写入数据
                row = grid_count + 1  # 数据从第2行开始
                ws[f"A{row}"] = f"Grid_{grid_count}"
                ws[f"B{row}"] = i
                ws[f"C{row}"] = j
                ws[f"D{row}"] = x1
                ws[f"E{row}"] = y1
                ws[f"F{row}"] = x2
                ws[f"G{row}"] = y2
                ws[f"H{row}"] = width
                ws[f"I{row}"] = height

                grid_count += 1

        # 自动调整列宽
        for col_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 2

        # 保存工作簿
        wb.save(output_path)
        print(f"网格位置信息已保存至: {output_path}")
        return output_path


def som_gridding(img, image_path):
    """调用MicroarrayGridding类进行网格划分"""
    gridding = MicroarrayGridding(image_path)
    return gridding.process()


def visualize_grid(img, adjusted_rows, adjusted_cols):
    """可视化网格划分结果"""
    visual_img = img.copy()
    for row in adjusted_rows:
        cv2.line(visual_img, (0, row), (visual_img.shape[1], row), (255, 255, 0), 1)
    for col in adjusted_cols:
        cv2.line(visual_img, (col, 0), (col, visual_img.shape[0]), (255, 255, 0), 1)
    return visual_img


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("自动网格化工具")

        self.image_path = None

        # 创建按钮
        self.select_button = tk.Button(root, text="选择图片", command=self.select_image)
        self.select_button.pack(pady=10)

        self.grid_button = tk.Button(root, text="自动网格化", command=self.auto_grid_action, state=tk.DISABLED)
        self.grid_button.pack(pady=10)

    def select_image(self):
        self.image_path = filedialog.askopenfilename(filetypes=[("图像文件", "*.png;*.jpg;*.jpeg")])
        if self.image_path:
            self.grid_button.config(state=tk.NORMAL)

    def auto_grid_action(self):
        if not self.image_path:
            print("请先选择图片")
            return

        # 加载图像
        img = cv2.imread(self.image_path)
        if img is None:
            print(f"无法加载图像: {self.image_path}")
            return

        # 执行网格划分
        gridding = MicroarrayGridding(self.image_path)
        adjusted_rows, adjusted_cols = gridding.process()

        # 可视化网格划分结果
        grid_visualization = visualize_grid(img, adjusted_rows, adjusted_cols)

        # 保存处理后的图像
        output_image_path = "gridded_result.png"  # 输出图像路径
        cv2.imwrite(output_image_path, grid_visualization)
        print(f"网格划分结果已保存至: {output_image_path}")

        # 保存网格位置信息到Excel
        excel_path = gridding.save_grid_to_excel(adjusted_rows, adjusted_cols)

        # 显示结果图像（可选）
        cv2.imshow("Grid Visualization", grid_visualization)
        cv2.waitKey(0)
        cv2.destroyAllWindows()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()